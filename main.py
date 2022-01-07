import time

import openpyxl
import xlrd

from locations import creat_location
from openpyxl import load_workbook


class Attach:
    def __init__(self, locator="baidu"):
        self.app = creat_location(locator)
        self.target_xls = "target.xlsx"
        self.work_book = load_workbook(self.target_xls)
        self.sheet = self.work_book.get_sheet_by_name("Sheet1")

    def load_data(self):
        workbook = xlrd.open_workbook(self.target_xls)
        sheet = workbook.sheet_by_index(0)
        address_list = sheet.col(1)[1:]
        for address in address_list:
            yield address.value

    def get_location(self, address: str):
        result_list = []
        split_lis = address.split('、')
        for item in split_lis:
            location = self.app.locate(item)
            result_list.append(location)
        return result_list

    def save_data(self, coordinates, rowx):
        if len(coordinates) == 1:
            self.sheet.cell(row=rowx + 2, column=3).value = str(coordinates[0])
            self.work_book.save(self.target_xls)
        else:
            self.sheet.cell(row=rowx + 2, column=3).value = str(coordinates).replace("[", "").replace("]", "")
            self.work_book.save(self.target_xls)

    @property
    def num_cell(self):
        workbook = xlrd.open_workbook(self.target_xls)
        sheet = workbook.sheet_by_index(0)
        return sheet.nrows - 1


attacher = Attach()
nrows = attacher.num_cell

for index, item in enumerate(attacher.load_data()):
    coordinate = attacher.get_location(item)
    attacher.save_data(coordinate, index)
    percent = (index + 1) / nrows * 100
    print(str(percent) + "%")
